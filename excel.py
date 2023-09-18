'''
Author       : facsert
Date         : 2023-08-14 10:43:49
LastEditTime: 2023-08-28 22:24:56
Description  : edit description
'''

from openpyxl import load_workbook, Workbook
from json import dumps
from toml import dump, load

class Excel:
    
    def __init__(self, file, mode):
        self.file = file
        self.mode = mode
        self.mode_init()
        
    def __enter__(self):
        return self
    
    def __exit__(self, e_type, e_value, e_tb):
        self.wb.close()
        if any((e_type, e_value, e_tb)):
            raise RuntimeError(f"error: {e_value}\n")
    
    def mode_init(self):
        if self.mode.lower() == 'w':
            self.wb = Workbook()
            self.sheet = self.wb.active
            return
            
        if self.mode.lower() == 'r':
            self.wb = load_workbook(self.file, data_only=True)
            self.sheet = self.wb[self.wb.sheetnames[0]]
            self.head = self.read_head()
            return 
        
        print(f'error mode {self.mode}, select r or w')
        exit()

    def cell_value(self, row, col):
        return self.sheet.cell(row=row, column=col).value
    
    def set_cell(self, row, col, value):
        self.sheet.cell(row, col).value = value
        
    def read_head(self):
        max_col = self.sheet.max_column + 1
        return [self.cell_value(1, col) for col in range(1, max_col)]
    
    def select_column(self, select, key=None):
        if len(select) == 0:
            select = self.head
            
        if key != None and key not in select:
            print(f"{key} not in {select}")
            exit()
        
        if set(self.head) < set(select):
            print(f'{select} not in {self.head}')
            exit()
        
        return [self.head.index(k) for k in select if k in self.head]
    
    def excel_to_list(self, select=[]):
        indexs = self.select_column(select)
        excel_list = []
        
        for row in range(2, self.sheet.max_row + 1):        
            excel_list.append({
                self.head[col]: self.cell_value(row, col+1) 
                for col in indexs
            })
        return excel_list
        
    def excel_to_dict(self, key, select=[]):
        indexs = self.select_column(select, key)
        excel_dict = {}
        
        for row in range(2, self.sheet.max_row + 1):
            key_value = self.cell_value(row, self.head.index(key) + 1)
            excel_dict.update({key_value: {
                self.head[col]: self.cell_value(row, col+1)
                for col in indexs
            }})
            
        return dict(sorted(excel_dict.items()))
    
    def list_to_excel(self, lst, select=[]):
        try:
            self.head = list(lst[0].keys())
        except Exception as e:
            print(f"data type error, must list[dict]")
            exit()
            
        indexs = self.select_column(select)
        self.sheet.append([self.head[i] for i in indexs])
        for line in lst:
            self.sheet.append([line[self.head[i]] for i in indexs])

        self.wb.save(self.file)
        
    def dict_to_excel(self, dic, select=[]):
        try:
            self.head = list(list(dic.values())[0].keys())
        except Exception as e:
            print(f"data type error, must dict[str, dict]")
            exit()
            
        indexs = self.select_column(select)
        self.sheet.append([self.head[i] for i in indexs])
        for line in dic.values():
            self.sheet.append([line[self.head[i]] for i in indexs])
            
        self.wb.save(self.file)


if __name__ == '__main__':
    with Excel('excel.xlsx', 'r') as e:
        array = e.excel_to_dict(key="用例编号", select=["用例编号", "用例名称", "执行结果"])
        
    with Excel('save.xlsx', 'w') as e:
        e.dict_to_excel(array, select=["用例编号", "执行结果"])